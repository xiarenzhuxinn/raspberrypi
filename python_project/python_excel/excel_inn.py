#-*-coding:utf-8-*-
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
file=load_workbook('test.xlsx')
sh1=file['Sheet1']
rows=sh1.max_row
cols=sh1.max_column
print(rows,cols)
for j in range(1,cols+1,5):
    with open('test.txt','r') as txt:
        for line in txt:
            row=int((len(line.strip('\n'))-1)/8)
            cp=line.strip('\n').split('\t')
            if sh1.cell(1,j).value==cp[0]:
                for t in range(0,4):
                    for r in range(2,row+1):
                        sh1.cell(r,j+t).value=cp[t+1+4*(r-2)]
file.save('test.xlsx')