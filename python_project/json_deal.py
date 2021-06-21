#!/usr/bin/python3
import json
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
file=load_workbook('test.xlsx')
sh1=file['Sheet1']
col=2

ljx = {"name":"桔子桑","sex":"男","age":18,"lljx":[{"ljx":"sd8081","ljxx":"11"},{"ljx":"hi1215","ljxx":"11"},{"ljx":"3","ljxx":"33"}]}
arg = ["ljx","ljxx"]
print(len(ljx["lljx"]))
for i in range(0,len(ljx)-1):
    if ljx["lljx"][i]["ljxx"] == "11":
        sh1.cell(1,i*col+1).value = ljx["lljx"][i]["ljx"]
        print(i)
file.save('test.xlsx')
